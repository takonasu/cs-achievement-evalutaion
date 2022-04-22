import openpyxl
import json
import collections as cl

load_book = openpyxl.load_workbook('ファイルパス')
sheet1 = load_book['Page 1']
sheet2 = load_book['Page 2']
sheet3 = load_book['Page 3']
sheet4 = load_book['Page 4']
json_path = '../data/point.json'


# 辞書作成
data_dict = cl.OrderedDict()

def get_data(sheetnum, sheetrow, sheetcol):
    sheet = sheet1
    if(sheetnum == 1):    
        sheet = sheet1
    elif(sheetnum == 2):    
        sheet = sheet2
    elif(sheetnum == 3):    
        sheet = sheet3
    elif(sheetnum == 4):    
        sheet = sheet4
    else:
        print("error sheet num")
        return
    
    # 科目名
    class_name = sheet.cell(row = sheetrow, column = sheetcol).value
    class_name_list = class_name.splitlines()
    ## 汎用コンピテンス(General Competence)
    class_g1 = sheet.cell(row = sheetrow, column = sheetcol+8).value
    class_g2 = sheet.cell(row = sheetrow, column = sheetcol+9).value
    class_g3 = sheet.cell(row = sheetrow, column = sheetcol+10).value
    class_g4 = sheet.cell(row = sheetrow, column = sheetcol+11).value
    class_g5 = sheet.cell(row = sheetrow, column = sheetcol+12).value
    ## 専門コンピテンス(Professional Competence)
    class_p1 = sheet.cell(row = sheetrow, column = sheetcol+13).value
    class_p2 = sheet.cell(row = sheetrow, column = sheetcol+14).value
    class_p3 = sheet.cell(row = sheetrow, column = sheetcol+15).value


    class_g1_list = class_g1.splitlines()
    class_g2_list = class_g2.splitlines()
    class_g3_list = class_g3.splitlines()
    class_g4_list = class_g4.splitlines()
    class_g5_list = class_g5.splitlines()
    
    class_p1_list = class_p1.splitlines()
    class_p2_list = class_p2.splitlines()
    class_p3_list = class_p3.splitlines()

    for i in range(len(class_name_list)):
        data_dict[class_name_list[i]] = cl.OrderedDict({"g1": int(class_g1_list[i]),"g2": int(class_g2_list[i]),"g3": int(class_g3_list[i]),"g4": int(class_g4_list[i]),"g5": int(class_g5_list[i]),"p1": int(class_p1_list[i]),"p2": int(class_p2_list[i]),"p3": int(class_p3_list[i])})

# データの取り出し

# 学術院共通専門基盤科目
get_data(1,4,4)

# 研究群共通科目
## 専門基礎科目
### リスク・レジリエンス工学関連科目
get_data(1,11,4)
 
### 情報理工学関連科目
get_data(1,13,4)


## 専門科目
### リスク・レジリエンス工学関連科目
get_data(2,14,5)
### 情報理工学関連科目
get_data(2,16,5)
get_data(2,17,5)
### 知能機能システム関連科目
get_data(2,19,5)
### 構造エネルギー工学関連科目
get_data(2,21,5)

## 学位プログラム科目群
# 専門基礎科目
get_data(4,18,5)
# 専門科目
get_data(4,20,5)

# 書き込み
with open(json_path, mode = 'w', encoding = 'utf-8') as f:
    f.write(json.dumps(data_dict, ensure_ascii = False, indent = 4))